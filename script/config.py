# encoding=utf-8
import os, csv
import openpyxl

WORK_PATH 	= 	os.getcwd()
CSV_FILE 	= 	""
EXCEL_FILE 	= 	"%s%s.xlsx" % (WORK_PATH, "\\sieges_excel")
MODE = ""

def generator_csv_file(mode):
	global CSV_FILE, MODE
	MODE = mode
	if os.name != "nt": #win	
		CSV_FILE	=	"%s%s%s.csv" % (WORK_PATH, "/sieges_output_", mode)
	else:
		CSV_FILE 	= 	"%s%s%s.csv" % (WORK_PATH, "\\sieges_output_", mode)
	#return CSV_FILE
	#print(CSV_FILE)

def writeExcel():
	wb = openpyxl.load_workbook('八年级学生基本信息.xlsx')
	for sheet_name in wb.sheetnames:
		sheet = wb[sheet_name]
		for i in range( 2, sheet.max_row + 1):
			yield int(sheet[i][8].value or 150)


'''
在Linux下安装matplotlib
1. 下载matploat 
2. 安装依赖freetype, png
'''

URL = "192.168.128.128/index.php"
#URL = "192.168.128.128"
X_REQ = [100, 200, 400, 800, 1000, 1200, 1400, 1600, 1800, 2000]
#X_REQ = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
#X_REQ = [100]
RECORD_ITEM = {"Response time": "请求时间(sec)", "Throughput": "吞吐量(KB/sec)"}
ARGS	=	["rr", "dyn", "ip"]

def write_output(datas):
	if not isinstance(datas, dict):
		raise TypeError("write_output need arg dict")
	headers = []
	for k in datas:
		headers.append(k)
	data = [datas]

	if not os.path.exists(CSV_FILE):
		with open(CSV_FILE, 'w', newline='') as f:
			writer = csv.DictWriter(f, headers)
			writer.writeheader()

	with open(CSV_FILE, 'a', newline='') as f:
		writer = csv.DictWriter(f, headers)
		writer.writerows(data)

def parse_output(data):
	if not isinstance(data, tuple):
		raise TypeError("parse_output need arg tuple")
	#print(data)
	dic = {}
	msg_arr = data[1].decode('utf-8').strip().split("\n")[-12:]
	
	for v in msg_arr:
		line_arr = v.split(":")
		dic[line_arr[0].strip()] = line_arr[1].strip()

	return dic

def read_csv_by_col(col_name):
	y_data = []
	with open(CSV_FILE,'r') as myFile:  
		lines = csv.DictReader(myFile)
		for line in lines:
			dd = float(line[col_name].split()[0])
			if col_name == "Throughput":
				dd = dd * 1000
			y_data.append(dd)
		data = {MODE: [X_REQ, y_data, col_name]}
		return data

def getCSV_FILE():
	return CSV_FILE

def write_excel(reader):
	from openpyxl import Workbook
	from openpyxl.chart import LineChart, Reference, Series
	from openpyxl.styles import Alignment
	align = Alignment(horizontal='center', vertical='center')
	wb = Workbook()
	ws = wb.active

	ws.merge_cells('A1:A2')
	ws['A1'].value = "并发量"
	for i, v in enumerate(X_REQ, 3):
		local = 'A%d' % i
		ws[local].value = v

	num_f = len(ARGS)
	num_req = len(X_REQ)
	start_chr = 'B'
	n = 0
	for item, val in RECORD_ITEM.items():
		start_col = 2 + num_f*n
		end_col = start_col + num_f - 1
		end_chr = chr(ord(start_chr) + num_f - 1)
		merge_cells = '%s1:%s1' % (start_chr, end_chr)
		ws.merge_cells(merge_cells)
		cell_item = "%s1" % start_chr
		ws[cell_item].value = val
		ws[cell_item].alignment = align

		
		cell_range_f = ws['%s2'%start_chr:'%s2'%end_chr]
		for i, func in enumerate(ARGS):
			cell_range_f[0][i].value = func
			cell_range_f[0][i].alignment = align

			generator_csv_file(func)
			try:
				data = reader.read_log(item)
			except FileNotFoundError as e:
				#print(e)
				continue
			#print(data)
			cell_range_data = ws["%s3"%cell_range_f[0][i].column:"%s%d"%(cell_range_f[0][i].column, 3+num_req-1)]
			for j, d in enumerate(data[func][1]):
				cell_range_data[j][0].value = d
				#print(cell_range_data[j][0])
				#
		chart = LineChart()
		chart.title = item     #图的标题

		chart.y_axis.title = val	#y坐标的标题
		chart.x_axis.title = "并发量"	#x坐标的标题
		labels = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=3+num_req-1)

		values = Reference(ws, min_col=start_col, min_row=2, max_col=end_col, max_row=3+num_req-1)
		chart.add_data(values, titles_from_data=True)
		pic_loc = "%s15" % chr(ord('A')+n*10)
		chart.set_categories(labels)
		ws.add_chart(chart, pic_loc)
	
		start_chr = chr(ord(end_chr) + 1)
		n += 1


	try:
		wb.save(EXCEL_FILE)
	except PermissionError as e:
		#print(e, "保存excel失败, 先关闭文件后重试")
		os.system('taskkill /IM excel.exe /F')
		wb.save(EXCEL_FILE)
	#print("正在自动打开文件, 请稍后...")
	#os.system(EXCEL_FILE)